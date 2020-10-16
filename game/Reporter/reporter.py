# -*- coding: utf-8 -*-

import openpyxl
import os



class Reporter_excel(object):

    def __init__(self):
        self.path = os.path.abspath(os.path.dirname(__file__)) + "//"
        self.template_path = self.path + 'report_template.xlsx'
        self.report_path = self.path + 'reporter.xlsx'
        self.title = '广告测试'
        self.isFirstOpen = True



    def read_excel(self):
        wb = openpyxl.load_workbook(self.template_path)
        sheet = wb[self.title]

        data = sheet['A1'].value
        print(data)

    def write_excel(self, cell=None, data=None):
        if self.isFirstOpen == True:
            wb = openpyxl.load_workbook(self.template_path)
            self.isFirstOpen = False
        else:
            wb = openpyxl.load_workbook(self.report_path)
        sheet = wb[self.title]
        sheet[cell] = data
        wb.save(self.report_path)

# b = Reporter()
# b.write_excel('B1', 'huawe')